"""
Apicbase Migration: Phase 2 - Import remaining data files.
Handles files with non-standard header structures.
"""
import asyncio
import json
import os
from datetime import datetime, timezone

from dotenv import load_dotenv
load_dotenv()

import openpyxl
from motor.motor_asyncio import AsyncIOMotorClient

MONGO_URL = os.getenv("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = "restin_v2"
DATA_DIR = os.path.join(os.path.dirname(__file__), "data", "apicbase_exports")


def safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(str(val).replace(",", "").replace("€", "").strip())
    except (ValueError, TypeError):
        return default


def read_sheet_smart(filepath, sheet_name=None, header_row=None):
    """Read Excel with smart header detection. Returns list of dicts."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    target = sheet_name or wb.sheetnames[0]
    if target not in wb.sheetnames:
        target = wb.sheetnames[0]
    ws = wb[target]

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(all_rows) < 2:
        return []

    # Find header row: first row where most cells (>50%) are non-empty strings
    if header_row is None:
        for i, row in enumerate(all_rows):
            non_empty = sum(1 for c in row if c is not None and str(c).strip())
            total = len([c for c in row if c is not None or True])
            if total > 0 and non_empty / max(len(row), 1) > 0.3:
                # Check if it looks like headers (mostly strings, no numbers)
                strings = sum(1 for c in row if c is not None and isinstance(c, str) and len(str(c).strip()) > 1)
                if strings >= 2:
                    header_row = i
                    break
        if header_row is None:
            header_row = 0

    headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(all_rows[header_row])]

    result = []
    for row in all_rows[header_row + 1:]:
        if not row or all(v is None for v in row):
            continue
        record = {}
        for i, val in enumerate(row):
            if i < len(headers):
                record[headers[i]] = val
        # Skip rows where first column is empty
        first_val = row[0] if row else None
        if first_val is not None and str(first_val).strip():
            result.append(record)

    return result


async def import_outlet_sell_prices(db):
    """Import outlet sell prices - header row is row 0."""
    print("\n=== OUTLET SELL PRICES ===")
    filepath = os.path.join(DATA_DIR, "outlet_sell_prices.xlsx")
    if not os.path.exists(filepath):
        print("  SKIP")
        return

    rows = read_sheet_smart(filepath, "Data Sheet")
    print(f"  Found {len(rows)} rows")

    docs = []
    for row in rows:
        recipe_name = str(row.get("recipe name", "")).strip()
        if not recipe_name:
            continue

        outlet_name = str(row.get("outlet name", "")).strip()
        venue_map = {
            "Caviar&Bull": "venue-caviar-bull",
            "Don Royale": "venue-don-royale",
            "Sole by Tarragon": "venue-sole-tarragon",
        }

        docs.append({
            "apicbase_outlet_id": str(row.get("outlet id", "")).strip(),
            "outlet_name": outlet_name,
            "venue_id": venue_map.get(outlet_name),
            "apicbase_recipe_id": str(row.get("recipe id", "")).strip(),
            "recipe_name": recipe_name,
            "sell_price": safe_float(row.get("sell price")),
            "tax_rate": safe_float(row.get("tax rate")),
            "target_profit_margin": safe_float(row.get("target profit margin")),
            "source": "apicbase",
            "created_at": datetime.now(timezone.utc),
        })

    if docs:
        await db.outlet_sell_prices.delete_many({"source": "apicbase"})
        result = await db.outlet_sell_prices.insert_many(docs)
        print(f"  Inserted {len(result.inserted_ids)} outlet sell prices")


async def import_pricebook(db):
    """Import pricebook - header row is row 0, simple structure."""
    print("\n=== PRICEBOOK ===")
    filepath = os.path.join(DATA_DIR, "pricebook_full.xlsx")
    if not os.path.exists(filepath):
        print("  SKIP")
        return

    rows = read_sheet_smart(filepath)
    print(f"  Found {len(rows)} rows")

    docs = []
    for row in rows:
        product = str(row.get("product", "")).strip()
        if not product:
            continue

        docs.append({
            "product_name": product,
            "stock_code": str(row.get("stockcode", "")).strip() or None,
            "measure": safe_float(row.get("measure", 1)),
            "unit": str(row.get("unit", "")).strip() or None,
            "price": safe_float(row.get("price")),
            "source": "apicbase",
            "created_at": datetime.now(timezone.utc),
        })

    if docs:
        await db.pricebook.delete_many({"source": "apicbase"})
        result = await db.pricebook.insert_many(docs)
        print(f"  Inserted {len(result.inserted_ids)} pricebook items")


async def import_procurement(db):
    """Import procurement - multi-sheet (one per supplier), headers on row 2."""
    print("\n=== PROCUREMENT ===")

    proc_files = [
        ("procurement_caviarbull.xlsx", "venue-caviar-bull", "Caviar&Bull"),
        ("procurement_donroyale.xlsx", "venue-don-royale", "Don Royale"),
    ]

    total = 0
    for filename, venue_id, outlet_name in proc_files:
        filepath = os.path.join(DATA_DIR, filename)
        if not os.path.exists(filepath):
            print(f"  SKIP: {filename}")
            continue

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            supplier_name = sn  # Each sheet is named after the supplier
            rows = read_sheet_smart(filepath, sn, header_row=2)

            docs = []
            for row in rows:
                ingredient = str(row.get("ingredient", "")).strip()
                if not ingredient:
                    continue

                docs.append({
                    "venue_id": venue_id,
                    "outlet_name": outlet_name,
                    "supplier_name": supplier_name,
                    "ingredient_name": ingredient,
                    "uid": str(row.get("uid", "")).strip() or None,
                    "stock_item": str(row.get("stock item", "")).strip() or None,
                    "accounting_category": str(row.get("acc. category", "")).strip() or None,
                    "quantity_ordered": safe_float(row.get("quantity ordered")),
                    "quantity_delivered": safe_float(row.get("quantity delivered")),
                    "actual_price": safe_float(row.get("actual price")),
                    "source": "apicbase",
                    "created_at": datetime.now(timezone.utc),
                })

            if docs:
                result = await db.procurement_history.insert_many(docs)
                total += len(result.inserted_ids)

        wb.close()
        print(f"  {filename}: imported from {len(wb.sheetnames)} supplier sheets")

    print(f"  Total procurement records: {total}")


async def import_waste(db):
    """Import waste - 'Waste Events Breakdown' sheet, headers on row 2."""
    print("\n=== WASTE ===")
    filepath = os.path.join(DATA_DIR, "waste_caviarbull.xlsx")
    if not os.path.exists(filepath):
        print("  SKIP")
        return

    rows = read_sheet_smart(filepath, "Waste Events Breakdown", header_row=2)
    print(f"  Found {len(rows)} waste event rows")

    docs = []
    for row in rows:
        event_id = str(row.get("event id", "")).strip()
        stock_item = str(row.get("stock item", "")).strip()
        if not stock_item:
            continue

        docs.append({
            "venue_id": "venue-caviar-bull",
            "outlet_name": "Caviar&Bull",
            "event_id": event_id,
            "event_name": str(row.get("event name", "")).strip() or None,
            "date": str(row.get("date", "")).strip() or None,
            "registered_by": str(row.get("registered by", "")).strip() or None,
            "stock_item": stock_item,
            "accounting_category": str(row.get("acc. category", "")).strip() or None,
            "waste_category": str(row.get("waste category", "")).strip() or None,
            "remarks": str(row.get("remarks", "")).strip() or None,
            "amount": safe_float(row.get("amount")),
            "value": safe_float(row.get("value")),
            "source": "apicbase",
            "created_at": datetime.now(timezone.utc),
        })

    if docs:
        await db.waste_history.delete_many({"source": "apicbase"})
        result = await db.waste_history.insert_many(docs)
        print(f"  Inserted {len(result.inserted_ids)} waste records")


async def import_sales(db):
    """Import sales data - headers on row 1."""
    print("\n=== SALES ===")
    filepath = os.path.join(DATA_DIR, "sales_caviarbull.xlsx")
    if not os.path.exists(filepath):
        print("  SKIP")
        return

    rows = read_sheet_smart(filepath, header_row=1)
    print(f"  Found {len(rows)} sales rows")

    docs = []
    for row in rows:
        product = str(row.get("product", "")).strip()
        if not product:
            continue

        docs.append({
            "venue_id": "venue-caviar-bull",
            "outlet_name": "Caviar&Bull",
            "product_name": product,
            "sales_amount": safe_float(row.get("sales amount")),
            "category_type": str(row.get("category type", "")).strip() or None,
            "source": "apicbase",
            "created_at": datetime.now(timezone.utc),
        })

    if docs:
        await db.sales_history.delete_many({"source": "apicbase"})
        result = await db.sales_history.insert_many(docs)
        print(f"  Inserted {len(result.inserted_ids)} sales records")


async def import_stock_count(db):
    """Import stock count - headers on row 2."""
    print("\n=== STOCK COUNT ===")
    filepath = os.path.join(DATA_DIR, "count_donroyale.xlsx")
    if not os.path.exists(filepath):
        print("  SKIP")
        return

    rows = read_sheet_smart(filepath, "Count report", header_row=2)
    print(f"  Found {len(rows)} stock count rows")

    docs = []
    for row in rows:
        ingredient = str(row.get("ingredient", "")).strip()
        if not ingredient:
            continue

        docs.append({
            "venue_id": "venue-don-royale",
            "outlet_name": "Don Royale",
            "ingredient_name": ingredient,
            "stock_item": str(row.get("stock item", "")).strip() or None,
            "uid": str(row.get("uid", "")).strip() or None,
            "accounting_category": str(row.get("acc. category", "")).strip() or None,
            "counted_quantity": safe_float(row.get("counted quantity")),
            "stock_value": safe_float(row.get("stock value")),
            "source": "apicbase",
            "created_at": datetime.now(timezone.utc),
        })

    if docs:
        await db.stock_counts.insert_many(docs)
        print(f"  Inserted {len(result.inserted_ids)} stock count records")


async def import_stock_list(db):
    """Import stock list for Caviar&Bull."""
    print("\n=== STOCK LIST ===")
    filepath = os.path.join(DATA_DIR, "stock_list_caviarbull.xlsx")
    if not os.path.exists(filepath):
        print("  SKIP")
        return

    rows = read_sheet_smart(filepath)
    print(f"  Found {len(rows)} stock list rows")

    docs = []
    for row in rows:
        name = None
        for key in ["ingredient", "name", "stock item", "product"]:
            if key in row and row[key]:
                name = str(row[key]).strip()
                break
        if not name:
            continue

        docs.append({
            "venue_id": "venue-caviar-bull",
            "outlet_name": "Caviar&Bull",
            "ingredient_name": name,
            "raw_data": {k: str(v)[:200] if v is not None else None for k, v in row.items()},
            "source": "apicbase",
            "created_at": datetime.now(timezone.utc),
        })

    if docs:
        await db.stock_snapshots.insert_many(docs)
        print(f"  Inserted {len(docs)} stock list records")


async def import_variance(db):
    """Import variance report for Don Royale."""
    print("\n=== VARIANCE ===")
    filepath = os.path.join(DATA_DIR, "variance_donroyale.xlsx")
    if not os.path.exists(filepath):
        print("  SKIP")
        return

    rows = read_sheet_smart(filepath, "Variance report", header_row=2)
    print(f"  Found {len(rows)} variance rows")

    docs = []
    for row in rows:
        ingredient = str(row.get("ingredient", "")).strip()
        if not ingredient:
            continue

        docs.append({
            "venue_id": "venue-don-royale",
            "outlet_name": "Don Royale",
            "ingredient_name": ingredient,
            "base_unit": str(row.get("base stock item unit", "")).strip() or None,
            "uid": str(row.get("uid", "")).strip() or None,
            "accounting_category": str(row.get("acc. category", "")).strip() or None,
            "actual_stock": safe_float(row.get("actual stock")),
            "theoretical_stock": safe_float(row.get("theoretical stock")),
            "stock_variance": safe_float(row.get("stock variance")),
            "actual_stock_value": safe_float(row.get("actual stock value")),
            "theoretical_stock_value": safe_float(row.get("theoretical stock value")),
            "stock_value_variance": safe_float(row.get("stock value variance")),
            "source": "apicbase",
            "created_at": datetime.now(timezone.utc),
        })

    if docs:
        await db.variance_reports.insert_many(docs)
        print(f"  Inserted {len(docs)} variance records")


async def final_summary(db):
    """Print final migration summary."""
    print("\n" + "=" * 60)
    print("FINAL MIGRATION SUMMARY (Phase 1 + 2)")
    print("=" * 60)

    collections = [
        "suppliers", "ingredients", "ingredient_packages",
        "outlet_sell_prices", "pricebook",
        "procurement_history", "waste_history", "sales_history",
        "stock_snapshots", "stock_counts", "variance_reports",
        "recipes_engineered", "recipes",
    ]

    for coll in collections:
        total = await db[coll].count_documents({})
        apic = await db[coll].count_documents({"source": "apicbase"})
        print(f"  {coll}: {total} total ({apic} from Apicbase)")

    print("=" * 60)


async def main():
    print("=" * 60)
    print("APICBASE -> RESTIN.AI MIGRATION PHASE 2")
    print(f"Started: {datetime.now()}")
    print("=" * 60)

    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]

    try:
        await import_outlet_sell_prices(db)
        await import_pricebook(db)
        await import_procurement(db)
        await import_waste(db)
        await import_sales(db)
        await import_stock_count(db)
        await import_stock_list(db)
        await import_variance(db)
        await final_summary(db)

        print(f"\nPhase 2 Completed: {datetime.now()}")
        print("All done!")

    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
    finally:
        client.close()


if __name__ == "__main__":
    asyncio.run(main())
