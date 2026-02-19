"""Phase 2b: Fix remaining imports (waste, stock count, stock list, variance)."""
import asyncio
import os
from datetime import datetime, timezone

from dotenv import load_dotenv
load_dotenv()

import openpyxl
from motor.motor_asyncio import AsyncIOMotorClient

MONGO_URL = os.getenv("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = "restin_v2"
DATA_DIR = os.path.join(os.path.dirname(__file__), "data", "apicbase_exports")


def safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(str(val).replace(",", "").replace("€", "").strip())
    except (ValueError, TypeError):
        return default


def read_sheet_fixed_header(filepath, sheet_name, header_row_idx):
    """Read Excel with a KNOWN header row index."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(all_rows) <= header_row_idx + 1:
        return []

    headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(all_rows[header_row_idx])]

    result = []
    for row in all_rows[header_row_idx + 1:]:
        if not row or all(v is None for v in row):
            continue
        if row[0] is None or str(row[0]).strip() == "":
            continue
        record = {}
        for i, val in enumerate(row):
            if i < len(headers):
                record[headers[i]] = val
        result.append(record)

    return result


async def fix_waste(db):
    """Waste Events Breakdown: headers are on row 2 (0-indexed)."""
    print("\n=== WASTE (FIX) ===")
    filepath = os.path.join(DATA_DIR, "waste_caviarbull.xlsx")
    if not os.path.exists(filepath):
        print("  SKIP")
        return

    rows = read_sheet_fixed_header(filepath, "Waste Events Breakdown", 2)
    print(f"  Found {len(rows)} waste event rows")

    if not rows:
        # Try sheet index debug
        wb = openpyxl.load_workbook(filepath, read_only=True)
        print(f"  Sheets: {wb.sheetnames}")
        for sn in wb.sheetnames:
            ws = wb[sn]
            rr = list(ws.iter_rows(max_row=4, values_only=True))
            for i, r in enumerate(rr):
                vals = [str(c)[:30] if c else "" for c in r[:5]]
                print(f"    {sn} R{i}: {vals}")
        wb.close()
        return

    docs = []
    for row in rows:
        stock_item = str(row.get("stock item", "")).strip()
        if not stock_item:
            continue

        docs.append({
            "venue_id": "venue-caviar-bull",
            "outlet_name": "Caviar&Bull",
            "event_id": str(row.get("event id", "")).strip(),
            "event_name": str(row.get("event name", "")).strip() or None,
            "date": str(row.get("date", "")).strip() or None,
            "registered_by": str(row.get("registered by", "")).strip() or None,
            "stock_item": stock_item,
            "accounting_category": str(row.get("acc. category", "")).strip() or None,
            "waste_category": str(row.get("waste category", "")).strip() or None,
            "remarks": str(row.get("remarks", "")).strip() or None,
            "amount": safe_float(row.get("amount")),
            "value": safe_float(row.get("value")),
            "source": "apicbase",
            "created_at": datetime.now(timezone.utc),
        })

    if docs:
        await db.waste_history.delete_many({"source": "apicbase"})
        result = await db.waste_history.insert_many(docs)
        print(f"  Inserted {len(result.inserted_ids)} waste records")


async def fix_stock_count(db):
    """Stock count: headers on row 2."""
    print("\n=== STOCK COUNT (FIX) ===")
    filepath = os.path.join(DATA_DIR, "count_donroyale.xlsx")
    if not os.path.exists(filepath):
        print("  SKIP")
        return

    rows = read_sheet_fixed_header(filepath, "Count report", 2)
    print(f"  Found {len(rows)} stock count rows")

    docs = []
    for row in rows:
        ingredient = str(row.get("ingredient", "")).strip()
        if not ingredient:
            continue

        docs.append({
            "venue_id": "venue-don-royale",
            "outlet_name": "Don Royale",
            "ingredient_name": ingredient,
            "stock_item": str(row.get("stock item", "")).strip() or None,
            "uid": str(row.get("uid", "")).strip() or None,
            "accounting_category": str(row.get("acc. category", "")).strip() or None,
            "counted_quantity": safe_float(row.get("counted quantity")),
            "stock_value": safe_float(row.get("stock value")),
            "source": "apicbase",
            "created_at": datetime.now(timezone.utc),
        })

    if docs:
        result = await db.stock_counts.insert_many(docs)
        print(f"  Inserted {len(result.inserted_ids)} stock count records")


async def fix_stock_list(db):
    """Stock list for Caviar&Bull."""
    print("\n=== STOCK LIST ===")
    filepath = os.path.join(DATA_DIR, "stock_list_caviarbull.xlsx")
    if not os.path.exists(filepath):
        print("  SKIP")
        return

    # Debug: show structure
    wb = openpyxl.load_workbook(filepath, read_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        rr = list(ws.iter_rows(max_row=4, values_only=True))
        print(f"  Sheet '{sn}' rows={ws.max_row}")
        for i, r in enumerate(rr):
            vals = [str(c)[:30] if c else "" for c in r[:8]]
            print(f"    R{i}: {vals}")
    wb.close()

    # Try reading from first visible sheet
    rows = read_sheet_fixed_header(filepath, wb.sheetnames[0], 0)
    if not rows:
        # Try header at row 1 or 2
        rows = read_sheet_fixed_header(filepath, wb.sheetnames[0], 1)
    if not rows:
        rows = read_sheet_fixed_header(filepath, wb.sheetnames[0], 2)

    print(f"  Found {len(rows)} stock records")

    docs = []
    for row in rows:
        name = None
        for key in ["ingredient", "name", "stock item", "product"]:
            if key in row and row[key]:
                name = str(row[key]).strip()
                break
        if not name:
            continue

        docs.append({
            "venue_id": "venue-caviar-bull",
            "outlet_name": "Caviar&Bull",
            "ingredient_name": name,
            "raw_data": {k: str(v)[:200] if v is not None else None for k, v in row.items()},
            "source": "apicbase",
            "created_at": datetime.now(timezone.utc),
        })

    if docs:
        result = await db.stock_snapshots.insert_many(docs)
        print(f"  Inserted {len(result.inserted_ids)} stock list records")


async def fix_variance(db):
    """Variance report for Don Royale: headers on row 2."""
    print("\n=== VARIANCE ===")
    filepath = os.path.join(DATA_DIR, "variance_donroyale.xlsx")
    if not os.path.exists(filepath):
        print("  SKIP")
        return

    rows = read_sheet_fixed_header(filepath, "Variance report", 2)
    print(f"  Found {len(rows)} variance rows")

    docs = []
    for row in rows:
        ingredient = str(row.get("ingredient", "")).strip()
        if not ingredient:
            continue

        docs.append({
            "venue_id": "venue-don-royale",
            "outlet_name": "Don Royale",
            "ingredient_name": ingredient,
            "base_unit": str(row.get("base stock item unit", "")).strip() or None,
            "uid": str(row.get("uid", "")).strip() or None,
            "accounting_category": str(row.get("acc. category", "")).strip() or None,
            "actual_stock": safe_float(row.get("actual stock")),
            "theoretical_stock": safe_float(row.get("theoretical stock")),
            "stock_variance": safe_float(row.get("stock variance")),
            "actual_stock_value": safe_float(row.get("actual stock value")),
            "theoretical_stock_value": safe_float(row.get("theoretical stock value")),
            "stock_value_variance": safe_float(row.get("stock value variance")),
            "source": "apicbase",
            "created_at": datetime.now(timezone.utc),
        })

    if docs:
        result = await db.variance_reports.insert_many(docs)
        print(f"  Inserted {len(docs)} variance records")


async def final_summary(db):
    """Print final summary."""
    print("\n" + "=" * 60)
    print("COMPLETE MIGRATION SUMMARY")
    print("=" * 60)

    collections = [
        "suppliers", "ingredients", "ingredient_packages",
        "outlet_sell_prices", "pricebook",
        "procurement_history", "waste_history", "sales_history",
        "stock_snapshots", "stock_counts", "variance_reports",
        "recipes_engineered", "recipes",
    ]

    for coll in collections:
        total = await db[coll].count_documents({})
        apic = await db[coll].count_documents({"source": "apicbase"})
        print(f"  {coll}: {total} total ({apic} from Apicbase)")

    print("=" * 60)


async def main():
    print("=" * 60)
    print("APICBASE MIGRATION PHASE 2b - FIXES")
    print(f"Started: {datetime.now()}")
    print("=" * 60)

    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]

    try:
        await fix_waste(db)
        await fix_stock_count(db)
        await fix_stock_list(db)
        await fix_variance(db)
        await final_summary(db)
        print(f"\nCompleted: {datetime.now()}")
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
    finally:
        client.close()


if __name__ == "__main__":
    asyncio.run(main())
