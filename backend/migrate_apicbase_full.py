"""
Full Apicbase → Restin.ai Migration Script
Reads all Excel exports from data/apicbase_exports/ and inserts into MongoDB.
Collections: ingredients, suppliers, ingredient_packages, outlet_sell_prices
"""
import asyncio
import json
import os
import sys
from datetime import datetime, timezone

from dotenv import load_dotenv
load_dotenv()

import openpyxl
from motor.motor_asyncio import AsyncIOMotorClient

MONGO_URL = os.getenv("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = "restin_v2"
DATA_DIR = os.path.join(os.path.dirname(__file__), "data", "apicbase_exports")

# Apicbase Outlet IDs → Restin venue_ids
VENUE_MAP = {
    "329600524550006": "venue-caviar-bull",
    "629400219190003": "venue-don-royale",
    "329600514510002": "venue-sole-tarragon",
    "Caviar&Bull": "venue-caviar-bull",
    "Don Royale": "venue-don-royale",
    "Sole by Tarragon": "venue-sole-tarragon",
}

# Will be populated from DB
VENUE_OID_MAP = {}

ALLERGEN_COLUMNS = [
    "corn", "wheat", "rye", "barley", "oats", "spelt", "khorasan",
    "crustaceans / crustacean shellfish", "eggs", "fish", "peanut",
    "cereals containing gluten", "soybeans", "milk", "lactose",
    "nuts / tree nuts", "walnuts", "pecan nuts", "brazil nuts",
    "pistachio nuts", "macadamia nuts", "almonds", "hazelnuts",
    "pine nuts", "chestnuts", "cashews", "celery", "mustard",
    "seeds", "sesame seeds", "poppy seeds", "sunflower seeds",
    "sulphites", "lupin", "molluscs",
]

DIETARY_COLUMNS = ["vegan", "vegetarian", "halal", "kosher"]

NUTRITION_COLUMNS = [
    "energy in kj", "energy in kcal", "fat (g)", "saturated fat (g)",
    "carbohydrate (g)", "sugar (g)", "fibre (g)", "protein (g)",
    "salt (g)", "sodium (g)",
]


def read_excel_rows(filepath, sheet_name=None):
    """Read an Excel file and return list of dicts (header->value)."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    target_sheet = sheet_name or "Data Sheet"
    if target_sheet not in wb.sheetnames:
        target_sheet = wb.sheetnames[0]
    ws = wb[target_sheet]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return []

    # Find the actual header row (first row with real column names)
    header_row_idx = 0
    for i, row in enumerate(rows):
        if row and row[0] and "apicbase" in str(row[0]).lower():
            header_row_idx = i
            break
        if row and row[0] and str(row[0]).strip() not in ("Internal", "General Information", "", None):
            header_row_idx = i
            break

    # If first row is category headers, real headers are row 1
    if rows[0] and rows[0][0] == "Internal":
        header_row_idx = 1

    headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(rows[header_row_idx])]

    result = []
    for row in rows[header_row_idx + 1:]:
        if not row or all(v is None for v in row):
            continue
        record = {}
        for i, val in enumerate(row):
            if i < len(headers):
                record[headers[i]] = val
        # Skip empty rows
        if record.get("apicbase id") or record.get("name") or record.get("name (required)"):
            result.append(record)

    return result


def safe_float(val, default=0.0):
    """Safely convert to float."""
    if val is None:
        return default
    try:
        return float(str(val).replace(",", "").replace("€", "").strip())
    except (ValueError, TypeError):
        return default


def safe_int(val, default=0):
    """Safely convert to int."""
    if val is None:
        return default
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return default


def extract_allergens(row):
    """Extract allergen data from a row."""
    allergens = {}
    for col in ALLERGEN_COLUMNS:
        val = row.get(col)
        if val is not None and val != "" and val != 0:
            allergens[col] = bool(int(val)) if str(val) in ("0", "1") else bool(val)
    return allergens


def extract_dietary(row):
    """Extract dietary flags from a row."""
    dietary = {}
    for col in DIETARY_COLUMNS:
        val = row.get(col)
        if val is not None and val != "" and val != 0:
            dietary[col] = bool(int(val)) if str(val) in ("0", "1") else bool(val)
    return dietary


def extract_nutrition(row):
    """Extract nutrition data from a row."""
    nutrition = {}
    for col in NUTRITION_COLUMNS:
        val = row.get(col)
        if val is not None and val != "":
            nutrition[col.split("(")[0].strip()] = safe_float(val)
    return nutrition


async def migrate_suppliers(db):
    """Import 106 suppliers from scraped JSON."""
    print("\n=== SUPPLIERS ===")
    json_path = os.path.join(DATA_DIR, "suppliers_scraped.json")
    if not os.path.exists(json_path):
        print("  SKIP: suppliers_scraped.json not found")
        return

    with open(json_path, "r", encoding="utf-8") as f:
        suppliers = json.load(f)

    # Also get supplier names from ingredient_packages for cross-reference
    pkg_path = os.path.join(DATA_DIR, "ingredient_packages.xlsx")
    pkg_suppliers = set()
    if os.path.exists(pkg_path):
        rows = read_excel_rows(pkg_path, "Packages List")
        for row in rows:
            sup_name = row.get("supplier")
            if sup_name and str(sup_name).strip():
                pkg_suppliers.add(str(sup_name).strip())

    # Merge: add any suppliers from packages not in scraped list
    existing_names = {s["name"] for s in suppliers}
    for pkg_sup in pkg_suppliers:
        if pkg_sup not in existing_names:
            suppliers.append({"name": pkg_sup, "type": "Supplier", "total_orders": "0"})

    docs = []
    for sup in suppliers:
        doc = {
            "name": sup["name"],
            "type": sup.get("type", "Supplier"),
            "contact_email": "",
            "contact_phone": "",
            "total_orders_value": safe_float(sup.get("total_orders", "0")),
            "source": "apicbase",
            "status": "active",
            "created_at": datetime.now(timezone.utc),
            "updated_at": datetime.now(timezone.utc),
        }
        docs.append(doc)

    if docs:
        # Clear existing apicbase suppliers
        await db.suppliers.delete_many({"source": "apicbase"})
        result = await db.suppliers.insert_many(docs)
        print(f"  Inserted {len(result.inserted_ids)} suppliers")
    return docs


async def migrate_ingredients(db):
    """Import ingredients from ingredient_list.xlsx."""
    print("\n=== INGREDIENTS ===")
    filepath = os.path.join(DATA_DIR, "ingredient_list.xlsx")
    if not os.path.exists(filepath):
        print("  SKIP: ingredient_list.xlsx not found")
        return

    rows = read_excel_rows(filepath)
    print(f"  Found {len(rows)} ingredient rows")

    docs = []
    for row in rows:
        apicbase_id = str(row.get("apicbase id", "")).strip()
        name = str(row.get("name (required)", row.get("name", ""))).strip()
        if not name:
            continue

        allergens = extract_allergens(row)
        dietary = extract_dietary(row)
        nutrition = extract_nutrition(row)

        doc = {
            "apicbase_id": apicbase_id,
            "name": name,
            "short_name": str(row.get("short name", "")).strip() or None,
            "internal_id": str(row.get("internal id", "")).strip() or None,
            "external_id": str(row.get("external id", "")).strip() or None,
            "brand": str(row.get("brand", "")).strip() or None,
            "article_type": str(row.get("article type", "")).strip() or None,
            "category": str(row.get("category", "")).strip() or None,
            "subcategory": str(row.get("subcategory", "")).strip() or None,
            "preferred_uom": str(row.get("preferred uom(s)", "")).strip() or None,
            "shelf_life": str(row.get("shelf life", "")).strip() or None,
            "storage_instructions": str(row.get("storage instructions", "")).strip() or None,
            "waste_percent": safe_float(row.get("waste %")),
            "mass_volume_ratio": safe_float(row.get("mass volume ratio")),
            "accounting_category": str(row.get("accounting category", "")).strip() or None,
            "verified": bool(safe_int(row.get("verified"))),
            "allowed": bool(safe_int(row.get("allowed"))),
            "allergens": allergens,
            "dietary": dietary,
            "nutrition": nutrition,
            "remarks": str(row.get("remarks / extra information", "")).strip() or None,
            "composition": str(row.get("composition", "")).strip() or None,
            "url": str(row.get("url", "")).strip() or None,
            "source": "apicbase",
            "status": "active",
            "created_at": datetime.now(timezone.utc),
            "updated_at": datetime.now(timezone.utc),
        }
        docs.append(doc)

    if docs:
        await db.ingredients.delete_many({"source": "apicbase"})
        result = await db.ingredients.insert_many(docs)
        print(f"  Inserted {len(result.inserted_ids)} ingredients")

    # Create indexes
    await db.ingredients.create_index("apicbase_id")
    await db.ingredients.create_index("name")
    await db.ingredients.create_index("category")
    return docs


async def migrate_ingredient_packages(db):
    """Import ingredient packages with supplier links."""
    print("\n=== INGREDIENT PACKAGES ===")
    filepath = os.path.join(DATA_DIR, "ingredient_packages.xlsx")
    if not os.path.exists(filepath):
        print("  SKIP: ingredient_packages.xlsx not found")
        return

    rows = read_excel_rows(filepath, "Packages List")
    print(f"  Found {len(rows)} package rows")

    docs = []
    for row in rows:
        apicbase_id = str(row.get("apicbase id", "")).strip()
        name = str(row.get("name", row.get("name (required)", ""))).strip()
        if not name:
            continue

        doc = {
            "apicbase_id": apicbase_id,
            "ingredient_name": name,
            "quantity": safe_float(row.get("quantity")),
            "unit": str(row.get("unit", "")).strip() or None,
            "gtin": str(row.get("gtin", "")).strip() or None,
            "package_type": str(row.get("package type", "")).strip() or None,
            "stockable": bool(safe_int(row.get("stockable"))),
            "piece": bool(safe_int(row.get("piece"))),
            "weighted": bool(safe_int(row.get("weighted"))),
            "supplier_name": str(row.get("supplier", "")).strip() or None,
            "supplier_article_number": str(row.get("supplier article #", "")).strip() or None,
            "product_name": str(row.get("product name", "")).strip() or None,
            "price": safe_float(row.get("price")),
            "price_per_pack": safe_float(row.get("price per pack")),
            "price_unit": str(row.get("price unit", "")).strip() or None,
            "orderable": bool(safe_int(row.get("orderable"))),
            "source": "apicbase",
            "created_at": datetime.now(timezone.utc),
        }
        docs.append(doc)

    if docs:
        await db.ingredient_packages.delete_many({"source": "apicbase"})
        result = await db.ingredient_packages.insert_many(docs)
        print(f"  Inserted {len(result.inserted_ids)} ingredient packages")

    await db.ingredient_packages.create_index("apicbase_id")
    await db.ingredient_packages.create_index("ingredient_name")
    return docs


async def migrate_outlet_sell_prices(db):
    """Import outlet-specific sell prices."""
    print("\n=== OUTLET SELL PRICES ===")
    filepath = os.path.join(DATA_DIR, "outlet_sell_prices.xlsx")
    if not os.path.exists(filepath):
        print("  SKIP: outlet_sell_prices.xlsx not found")
        return

    rows = read_excel_rows(filepath, "Data Sheet")
    print(f"  Found {len(rows)} sell price rows")

    docs = []
    for row in rows:
        outlet_id = str(row.get("outlet id", "")).strip()
        recipe_id = str(row.get("recipe id", "")).strip()
        recipe_name = str(row.get("recipe name", "")).strip()
        if not recipe_name:
            continue

        outlet_name = str(row.get("outlet name", "")).strip() or None
        venue_id = VENUE_MAP.get(outlet_id, VENUE_MAP.get(outlet_name))

        doc = {
            "apicbase_outlet_id": outlet_id,
            "outlet_name": outlet_name,
            "venue_id": venue_id,
            "apicbase_recipe_id": recipe_id,
            "recipe_name": recipe_name,
            "sell_price": safe_float(row.get("sell price")),
            "tax_rate": safe_float(row.get("tax rate")),
            "target_profit_margin": safe_float(row.get("target profit margin")),
            "source": "apicbase",
            "created_at": datetime.now(timezone.utc),
        }
        docs.append(doc)

    if docs:
        await db.outlet_sell_prices.delete_many({"source": "apicbase"})
        result = await db.outlet_sell_prices.insert_many(docs)
        print(f"  Inserted {len(result.inserted_ids)} outlet sell prices")
    return docs


async def migrate_stock_data(db):
    """Import stock data from stock reports."""
    print("\n=== STOCK DATA ===")

    stock_files = [
        ("stock_caviarbull.xlsx", "venue-caviar-bull", "Caviar&Bull"),
        ("stock_list_caviarbull.xlsx", "venue-caviar-bull", "Caviar&Bull"),
    ]

    total = 0
    for filename, venue_id, outlet_name in stock_files:
        filepath = os.path.join(DATA_DIR, filename)
        if not os.path.exists(filepath):
            print(f"  SKIP: {filename} not found")
            continue

        rows = read_excel_rows(filepath)
        print(f"  {filename}: {len(rows)} rows")

        docs = []
        for row in rows:
            name = str(row.get("name", row.get("ingredient", row.get("product", "")))).strip()
            if not name:
                continue

            doc = {
                "ingredient_name": name,
                "venue_id": venue_id,
                "outlet_name": outlet_name,
                "quantity": safe_float(row.get("quantity", row.get("stock", row.get("current stock")))),
                "unit": str(row.get("unit", row.get("uom", ""))).strip() or None,
                "value": safe_float(row.get("value", row.get("stock value"))),
                "category": str(row.get("category", "")).strip() or None,
                "storage_location": str(row.get("storage location", row.get("location", ""))).strip() or None,
                "source": "apicbase",
                "snapshot_date": datetime.now(timezone.utc),
                "created_at": datetime.now(timezone.utc),
            }
            docs.append(doc)

        if docs:
            result = await db.stock_snapshots.insert_many(docs)
            total += len(result.inserted_ids)

    print(f"  Total stock records: {total}")


async def migrate_procurement(db):
    """Import procurement data."""
    print("\n=== PROCUREMENT ===")

    proc_files = [
        ("procurement_caviarbull.xlsx", "venue-caviar-bull", "Caviar&Bull"),
        ("procurement_donroyale.xlsx", "venue-don-royale", "Don Royale"),
    ]

    total = 0
    for filename, venue_id, outlet_name in proc_files:
        filepath = os.path.join(DATA_DIR, filename)
        if not os.path.exists(filepath):
            print(f"  SKIP: {filename} not found")
            continue

        rows = read_excel_rows(filepath)
        print(f"  {filename}: {len(rows)} rows")

        docs = []
        for row in rows:
            # Procurement files have varying structures, capture all fields
            doc = {
                "venue_id": venue_id,
                "outlet_name": outlet_name,
                "raw_data": {k: str(v) if v is not None else None for k, v in row.items()},
                "source": "apicbase",
                "created_at": datetime.now(timezone.utc),
            }

            # Try to extract key fields
            for key in ["supplier", "supplier name"]:
                if key in row and row[key]:
                    doc["supplier_name"] = str(row[key]).strip()
                    break

            for key in ["total", "amount", "total amount", "value"]:
                if key in row and row[key]:
                    doc["amount"] = safe_float(row[key])
                    break

            for key in ["date", "order date", "delivery date"]:
                if key in row and row[key]:
                    doc["date"] = str(row[key]).strip()
                    break

            docs.append(doc)

        if docs:
            result = await db.procurement_history.insert_many(docs)
            total += len(result.inserted_ids)

    print(f"  Total procurement records: {total}")


async def migrate_waste(db):
    """Import waste data."""
    print("\n=== WASTE ===")
    filepath = os.path.join(DATA_DIR, "waste_caviarbull.xlsx")
    if not os.path.exists(filepath):
        print("  SKIP: waste_caviarbull.xlsx not found")
        return

    rows = read_excel_rows(filepath)
    print(f"  Found {len(rows)} waste rows")

    docs = []
    for row in rows:
        doc = {
            "venue_id": "venue-caviar-bull",
            "outlet_name": "Caviar&Bull",
            "raw_data": {k: str(v) if v is not None else None for k, v in row.items()},
            "source": "apicbase",
            "created_at": datetime.now(timezone.utc),
        }
        docs.append(doc)

    if docs:
        result = await db.waste_history.insert_many(docs)
        print(f"  Inserted {len(result.inserted_ids)} waste records")


async def migrate_pricebook(db):
    """Import pricebook data."""
    print("\n=== PRICEBOOK ===")
    filepath = os.path.join(DATA_DIR, "pricebook_full.xlsx")
    if not os.path.exists(filepath):
        print("  SKIP: pricebook_full.xlsx not found")
        return

    rows = read_excel_rows(filepath)
    print(f"  Found {len(rows)} pricebook rows")

    docs = []
    for row in rows:
        product = str(row.get("product", "")).strip()
        if not product:
            continue

        doc = {
            "product_name": product,
            "stock_code": str(row.get("stockcode", "")).strip() or None,
            "measure": safe_float(row.get("measure", 1)),
            "unit": str(row.get("unit", "")).strip() or None,
            "price": safe_float(row.get("price")),
            "source": "apicbase",
            "created_at": datetime.now(timezone.utc),
        }
        docs.append(doc)

    if docs:
        await db.pricebook.delete_many({"source": "apicbase"})
        result = await db.pricebook.insert_many(docs)
        print(f"  Inserted {len(result.inserted_ids)} pricebook items")


async def update_venue_ids(db):
    """Map venue string IDs to actual MongoDB ObjectIds."""
    print("\n=== VENUE MAPPING ===")
    async for v in db.venues.find({}, {"name": 1}):
        name = v.get("name", "")
        oid = str(v["_id"])
        if "Caviar" in name:
            VENUE_OID_MAP["venue-caviar-bull"] = oid
        elif "Don Royale" in name:
            VENUE_OID_MAP["venue-don-royale"] = oid
        elif "Sole" in name or "Tarragon" in name:
            VENUE_OID_MAP["venue-sole-tarragon"] = oid
        print(f"  {name} -> {oid}")


async def print_summary(db):
    """Print final migration summary."""
    print("\n" + "=" * 60)
    print("MIGRATION SUMMARY")
    print("=" * 60)

    collections = [
        "suppliers", "ingredients", "ingredient_packages",
        "outlet_sell_prices", "stock_snapshots",
        "procurement_history", "waste_history", "pricebook",
        "recipes_engineered", "recipes",
    ]

    for coll in collections:
        count = await db[coll].count_documents({})
        apic_count = await db[coll].count_documents({"source": "apicbase"})
        print(f"  {coll}: {count} total ({apic_count} from Apicbase)")

    print("=" * 60)


async def main():
    print("=" * 60)
    print("APICBASE -> RESTIN.AI FULL MIGRATION")
    print(f"Data directory: {DATA_DIR}")
    print(f"MongoDB: {MONGO_URL}")
    print(f"Database: {DB_NAME}")
    print(f"Started: {datetime.now()}")
    print("=" * 60)

    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]

    try:
        await update_venue_ids(db)
        await migrate_suppliers(db)
        await migrate_ingredients(db)
        await migrate_ingredient_packages(db)
        await migrate_outlet_sell_prices(db)
        await migrate_stock_data(db)
        await migrate_procurement(db)
        await migrate_waste(db)
        await migrate_pricebook(db)
        await print_summary(db)

        print(f"\nCompleted: {datetime.now()}")
        print("Migration finished successfully!")

    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        client.close()


if __name__ == "__main__":
    asyncio.run(main())
